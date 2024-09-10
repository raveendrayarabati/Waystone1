import re
import time
import openpyxl
import requests
from lxml import html
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import streamlit as st
from datetime import datetime, timedelta
from io import BytesIO
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode

from src.cusip_to_ticker import parallel_fetch_tickers, map_sec13f, map_eod

from src.FactSet import FormulaDataProcessor
import concurrent.futures



def get_input(key_prefix):
    """Helper function to get top-left, bottom-right cell coordinates, and header option from the user."""
    st.markdown(
        ":blue[Please choose the coordinates for the rectangular selection that corresponds to the :green[\"CUSIP\"] first value and :green[\"Quantity\"] first value columns in the client's Excel sheet.]")
    col1, col2 = st.columns(2)
    with col1:
        top_left = st.text_input("Top-Left Cell (e.g., A1)", key=f"{key_prefix}_top_left")
    with col2:
        bottom_right = st.text_input("Bottom-Right Cell (e.g., B10)", key=f"{key_prefix}_bottom_right")

    header = st.checkbox("Use first row as header", key=f"{key_prefix}_header", value=True)
    # header = 0 if use_header else None
    return top_left, bottom_right, header


def load_dataframe_from_excel(file, top_left, bottom_right, header):
    """Loads a DataFrame from an Excel file based on specified cell range and header option."""
    if top_left and bottom_right:
        # Adjust 'skiprows' based on whether the first row is used as a header
        skiprows = int(re.findall(r'\d+', top_left)[0]) - (2 if header else 1)
        usecols = f"{top_left[0].upper()}:{bottom_right[0].upper()}"  # Columns to use

        # Load the DataFrame with or without the header
        if header:
            df = pd.read_excel(file, usecols=usecols, skiprows=skiprows, header=0)
        else:
            df = pd.read_excel(file, usecols=usecols, skiprows=skiprows, header=None)

        return df.dropna(how='all')  # Drop all empty rows
    else:
        # Default behavior when no coordinates are provided
        return pd.read_excel(file, header=0 if header else None).dropna(how='all')


def handle_ticker():
    print('Mapping client Tickers')


def map_eod_sec13f(client_data_df, sec13f_df, cusip_col, quantity_col):
    # Map Cusip to Ticker from QuantumOnline with parallelization
    client_ticker_mapped_df = parallel_fetch_tickers(client_data_df, 'Cusip')
    print(client_ticker_mapped_df.shape)


def record_function_runtime(func):
    """
    A decorator that records the start and end times of a function's execution and prints the runtime in a 1m20s format.

    Parameters:
    - func: The function to be wrapped.

    Returns:
    - wrapper: The wrapped function with timing functionality.
    """

    def wrapper(*args, **kwargs):
        start_time = time.time()  # Record start time
        result = func(*args, **kwargs)  # Execute the function
        end_time = time.time()  # Record end time
        runtime_seconds = int(end_time - start_time)  # Calculate runtime in seconds
        minutes, seconds = divmod(runtime_seconds, 60)  # Convert to minutes and seconds
        runtime_formatted = f"{minutes}m {seconds}s"  # Format the runtime
        print(f"Function {func.__name__} took {runtime_formatted} to complete.")
        return result, runtime_formatted

    return wrapper


# @record_function_runtime
# def run_mappings(client_df: pd.DataFrame,
#                  sec13f_df: pd.DataFrame,
#                  eod_df: pd.DataFrame,
#                  cusip_col: str,
#                  quantity_col: str):
#     start_time = time.time()  # Record start time
#     client_ticker_mapped_df = parallel_fetch_tickers(client_df, cusip_col=cusip_col)
#     print(client_ticker_mapped_df.shape)


#     R1 = map_sec13f(client_ticker_mapped_df, sec13f_df, cusip_col=cusip_col)
#     R2 = map_eod(R1, eod_df)


#     end_time = time.time()  # Record end time
#     runtime_seconds = int(end_time - start_time)  # Calculate runtime in seconds
#     minutes, seconds = divmod(runtime_seconds, 60)  # Convert to minutes and seconds
#     runtime_formatted = f"{minutes}m {seconds}s"  # Format the runtime
#     return R2, runtime_formatted


def merge_results(client_df: pd.DataFrame,
                  cusip_col: str,
                  result_df: pd.DataFrame,
                  sec13f_df: pd.DataFrame):
    
    #deleted status based...
    sec13f_df = sec13f_df[sec13f_df['STATUS'].str.strip() != 'DELETED']
    sec13f_df.reset_index(drop=True, inplace=True)
    # Sanitize columns...
    client_df[cusip_col] = client_df[cusip_col].str.replace(' ', '')
    sec13f_df['CUSIP NO_formatted'] = sec13f_df['CUSIP NO'].str.replace(' ', '')

    result_df = result_df.drop(['Country', 'fsymId'], axis=1, errors='ignore')

    merged_df = pd.merge(client_df, sec13f_df, left_on=cusip_col, right_on='CUSIP NO_formatted', how='left',
                         suffixes=('', '_drop'))
    merged_df = merged_df.loc[:, ~merged_df.columns.str.endswith('_drop')]
    merged_df = merged_df.drop(['CUSIP NO_formatted', 'ASTRK'], axis=1, errors='ignore')
    result_df = result_df.drop_duplicates(subset = ["requestId","CUSIP","EODPrice","Ticker"],keep='first')
    R = merged_df.merge(result_df, left_on=cusip_col, right_on='requestId', how='left', suffixes=('', '_drop'))
    # R = R.drop_duplicates().reset_index()
    # R.drop("index", axis=1, inplace=True)
    R = R.loc[:, ~R.columns.str.endswith('_drop')]

    R['Ticker'] = R.apply(lambda x: str(x['Ticker']).replace('-USA', ''), axis=1)
    R['Ticker'] = R.apply(lambda x: None if str(x['Ticker']) == 'None' else x['Ticker'], axis=1)
    R = R.rename({'CUSIP NO': 'CUSIP (SEC)', 'ISSUER NAME': 'Issuer Name (SEC)', 'ISSUER DESCRIPTION': 'Class (SEC)',
                  'STATUS': 'Status (SEC)'}, axis=1)
    R = R.drop(['requestId', 'CUSIP', 'ASTRK'], axis=1, errors='ignore')
    # R = R.drop(['', ''], axis=1, errors='ignore')
    print(R.head())
    # R = R.drop(['CUSIP'], axis=1, errors='ignore')
    return R

def run_mappings(client_df: pd.DataFrame,
                 sec13f_df: pd.DataFrame,
                 cusip_col: str,
                 quantity_col: str,
                 price_as_on_date: str):
    fs_processor = FormulaDataProcessor()

    # Gather Identifiers from client_df
    ids = [str(x).strip() for x in client_df[cusip_col].tolist()]
    # ids = ["252131107", "38141G104", "457669AB5", "501889208", "61174X109", "759916AC3", "V7780T103", "AAPL"]
    cross_formulas = ["FF_CUSIP(CURR)", f"FG_PRICE({price_as_on_date})", "FSYM_TICKER_EXCHANGE", "P_EXCOUNTRY"]
    timeSeries_formulas = ["FF_CUSIP(CURR)", "P_PRICE(NOW)", "FSYM_TICKER_EXCHANGE", "P_EXCOUNTRY"]
    display_names = ["CUSIP", "EODPrice", "Ticker", "Country"]

    start_time = time.time()  # Record start time
    cross_series_df = fs_processor.fetch_data(ids, cross_formulas, display_names)
    time.sleep(3)
    time_Series_df = fs_processor.fetch_time_series_data(ids, timeSeries_formulas, display_names)
    # cross_series_df=fetch_with_retry(fs_processor.fetch_data(ids, cross_formulas, display_names))
    # time_Series_df =fetch_with_retry(fs_processor.fetch_time_series_data(ids, timeSeries_formulas, display_names))
    merged_df = pd.merge(cross_series_df, time_Series_df, on='requestId', suffixes=('_df1', '_df2'))
    cross_series_df['EODPrice'].fillna(merged_df['EODPrice_df2'], inplace=True)
    cross_series_df['Ticker'].fillna(merged_df['Ticker_df2'], inplace=True)

    end_time = time.time()  # Record end time

    runtime_seconds = int(end_time - start_time)  # Calculate runtime in seconds
    minutes, seconds = divmod(runtime_seconds, 60)  # Convert to minutes and seconds
    runtime_formatted = f"{minutes}m {seconds}s"  # Format the runtime

    result_df = cross_series_df

    if result_df is not None:
        try:
            # st.dataframe(results_df)
            st.info(f"Mapping client data from FactSet API took {runtime_formatted}")
            R = merge_results(client_df, cusip_col=cusip_col, result_df=result_df, sec13f_df=sec13f_df)
            return R
        except Exception as e:
            st.error(f"Exception caught: {str(e)}")


def check_client_df_sanity(df: pd.DataFrame, cusip_col: str, quantity_col: str):
    # Check if the DataFrame has exactly 2 columns
    # if len(df.columns) != 2:
    #     return False, "Please ensure that the first sheet of your workbook contains only two columns \"Identifier\" and \"Quantity\". "

    # Check if the column names match the expected cusip_col and quantity_col
    if not set([cusip_col, quantity_col]).issubset(df.columns):
        return False, f"Error: DataFrame does not have the specified columns '{cusip_col}' and '{quantity_col}'."

    # Check for non-empty values in the cusip_col
    if df[cusip_col].isnull().any() or (df[cusip_col] == '').any():
        return False, f"Error: '{cusip_col}' column contains empty or missing values."

    # Convert the column to strings
    df[cusip_col] = df[cusip_col].astype(str)

    # Add check for cusip_col values to be 9-digit alphanumeric strings only
    if not df[cusip_col].str.match(r'^[A-Za-z0-9]{9}$').all():
        return False, f"Error: '{cusip_col}' column must contain 9-digit alphanumeric strings only."

    # Check for non-null values in the quantity_col and parsability as int or float
    if not all(df[quantity_col].apply(lambda x: isinstance(x, (int, float)) and not pd.isnull(x))):
        return False, f"Error: '{quantity_col}' column contains null values or values that cannot be parsed as int or float."

    # If all checks pass
    return True, "Uploaded client data is in expected format."


def insert_dataframe_into_worksheet(ws, df, start_row=5, start_col=1):
    """
    Insert a pandas DataFrame into an existing openpyxl worksheet starting from
    a specific row and column.


    Parameters:
    - ws: The openpyxl worksheet instance where the DataFrame will be written.
    - df: The pandas DataFrame to write.
    - start_row: The starting row index in the worksheet where the DataFrame will begin (1-based).
    - start_col: The starting column index in the worksheet where the DataFrame will begin (1-based).
    """
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)


def write_13f_excel(sec_13f_df,client_data_file_name):
    # Create a new Excel workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = 'SEC 13F Report'
    ws['A2'] = client_data_file_name 
    from datetime import datetime
    # today_date = datetime.now().strftime('%Y-%m-%d')
    # Get the current date
    current_date = datetime.now()
    # Calculate the last day of the previous quarter
    default_as_on_date = last_date_of_previous_quarter()
    ws['A3'] = f'As Of: {default_as_on_date}'
    # ws['A2'] = eod_date
    count = len(sec_13f_df)
    total_market_value = sec_13f_df['Market Value'].sum()

    count_row = len(sec_13f_df) + 10
    total_market_value_row = count_row + 1

    # Insert Count and Total Market Value
    ws[f'A{count_row}'] = 'Count:'
    ws[f'B{count_row}'] = count
    ws[f'A{total_market_value_row}'] = 'Total Market Value:'
    ws[f'B{total_market_value_row}'] = total_market_value

    # Insert the DataFrame into the worksheet starting from row 5
    insert_dataframe_into_worksheet(ws, sec_13f_df, start_row=5)

    # Save the workbook to a binary object in memory
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)  # Rewind the buffer

    return excel_file


def convert_ws_13f(ws_df):
    sec_13f_report_df = ws_df.copy()
    sec_13f_report_df = sec_13f_report_df.rename({'Issuer Name (SEC)': 'Security Name'}, axis=1)
    sec_13f_report_df = sec_13f_report_df.rename({'Class (SEC)': 'Class'}, axis=1)
    sec_13f_report_df = sec_13f_report_df.rename({'CUSIP (Client)': 'CUSIP'}, axis=1)
    sec_13f_report_df['FIGI'] = ''
    sec_13f_report_df = sec_13f_report_df.rename({'Market Value (Quantity*Price)': 'Market Value'}, axis=1)
    sec_13f_report_df = sec_13f_report_df.rename({'Quantity': 'Shares'}, axis=1)
    sec_13f_report_df = sec_13f_report_df.rename({'Discretion Type': 'Discretion'}, axis=1)
    sec_13f_report_df = sec_13f_report_df.rename({'Other Managers': 'Managers'}, axis=1)

    sec_13f_report_df = sec_13f_report_df.drop(
        ['Ticker (Client)', 'CUSIP (SEC)', 'Match?', 'Price', 'De Minimis?', 'Complete?'], axis=1, errors='ignore')

    sec_13f_report_df = sec_13f_report_df[
        ['Security Name', 'Class', 'CUSIP', 'FIGI', 'Market Value', 'Shares', 'SH/PRN', 'PUT/CALL', 'Discretion',
         'Managers', 'Sole', 'Shared', 'None']]

    return sec_13f_report_df


def generate_ws_13f(ws_revised=None,client_data_file_name=None):
    if ws_revised is None and 'working_sheet_df' in st.session_state:
        ws_df = st.session_state['working_sheet_df']
        sec_13f_df = convert_ws_13f(ws_df)
    elif ws_revised is not None:
        sec_13f_df = convert_ws_13f(ws_revised)
    else:
        st.error("Please Generate a Working sheet before attempting to download 13F (Ecxel)")
    return write_13f_excel(sec_13f_df,client_data_file_name)


def generate_13f_from_ws(ws_df,client_data_file_name):
    sec_13f_excel = generate_ws_13f(ws_df,client_data_file_name)
    return sec_13f_excel
    sec_13f_cols = ['Security Name', 'Class', 'CUSIP', 'FIGI', 'Market Value', 'Shares', 'SH/PRN', 'PUT/CALL',
                    'Discretion', 'Managers', 'Sole', 'Shared', 'None']

    # Create an empty DataFrame with the specified columns
    sec_13f = pd.DataFrame(columns=sec_13f_cols)

    # Create an in-memory bytes buffer
    excel_io = BytesIO()

    # Use the ExcelWriter to write the DataFrame to the buffer as an Excel file
    with pd.ExcelWriter(excel_io, engine='xlsxwriter') as writer:
        sec_13f.to_excel(writer, index=False, sheet_name='13F')

    # Return the in-memory buffer containing the Excel file
    excel_io.seek(0)  # Go to the start of the stream
    return excel_io

def parallel_fetch_cusips(client_df, ticker_col, cusip_col):
    def fetch_cusip_from_tickers(tickers):
        processor = FormulaDataProcessor()
        ids = tickers
        display_names = ["cusip"]
        timeSeries_formulas = ["FF_CUSIP(CURR)"]
        time_series_df = processor.fetch_time_series_data(ids, timeSeries_formulas, display_names)
        return time_series_df

    def extract_ticker(ticker_value):
        # Ensure the ticker_value is a string
        ticker_str = str(ticker_value)
        # Split the string by both spaces and underscores
        components = re.split(r'[ _.]', ticker_str)
        # Assume the ticker is always the first part
        return components[0]

    # Collect all tickers that need CUSIP fetching
    tickers_to_fetch = []
    for index, row in client_df.iterrows():
        if pd.isnull(row[cusip_col]) or row[cusip_col] == "":
            ticker = extract_ticker(row[ticker_col])
            tickers_to_fetch.append(ticker)

    # Fetch CUSIPs for all tickers in one go
    if tickers_to_fetch:
        cusip_dict = fetch_cusip_from_tickers(tickers_to_fetch)
        # Update client_df with fetched CUSIPs
        for index, row in client_df.iterrows():
            # sleep(0.01)
            if pd.isnull(row[cusip_col]) or row[cusip_col] == "":
                ticker = extract_ticker(row[ticker_col])
                for index1, row1 in cusip_dict.iterrows():
                    if ticker == row1["requestId"]:
                        client_df.loc[index, cusip_col] = cusip_dict.loc[index1, "cusip"]

def last_date_of_previous_quarter():
    # Get the current date
    current_date = datetime.now()

    # Determine the current quarter
    current_quarter = (current_date.month - 1) // 3 + 1

    # Calculate the first month of the current quarter
    first_month_current_quarter = (current_quarter - 1) * 3 + 1

    # Calculate the first day of the current quarter
    first_day_current_quarter = datetime(current_date.year, first_month_current_quarter, 1)

    # Calculate the last day of the previous quarter
    last_day_previous_quarter = first_day_current_quarter - timedelta(days=1)

    return last_day_previous_quarter.strftime('%Y-%m-%d')

def Cleaning_Top_And_Bottom_rows(df,combined_df):
    unnamed_columns = df.columns[df.columns.str.startswith('Unnamed:') & ~df.columns.isna()]

    percentage_unnamed = len(unnamed_columns) / len(df.columns)

    # Check if more than 70% of columns start with 'Unnamed:'
    if percentage_unnamed > 0.7:
        # Replace column names with the first row values
        df.columns = df.iloc[0]  # .fillna(df.columns)
        # Drop the first row
        df = df.drop(df.index[0])

    nan_count = df.columns.isnull().sum()
    total_columns = len(df.columns)
    nan_proportion = nan_count / total_columns

    # Check if more than half of the column names are NaN
    if nan_proportion > 0.7:
        while any(pd.isna(df.columns)):
            # Replace column names with the first row values
            df.columns = df.iloc[0]  # .fillna(df.columns)
            # Drop the first row
            df = df.drop(df.index[0])

    # Reset the index if needed
    df.reset_index(drop=True, inplace=True)

    # Function to filter last 20 rows based on NaN percentage
    last_20_rows = df.tail(20)

    # Count None values in each of the last 10 rows
    none_counts = last_20_rows.isna().sum(axis=1)

    # Calculate the threshold for None values
    none_threshold = 0.6 * len(df.columns)

    # Filter rows where None values exceed the threshold
    filtered_rows = last_20_rows[none_counts >= none_threshold]

    # Drop the identified rows from the original DataFrame
    df.drop(filtered_rows.index, inplace=True)

    #
    df = df.reset_index(drop=True)
    combined_df = combined_df.reset_index(drop=True)
    print(df)
    combined_df = pd.concat([combined_df, df], ignore_index=True, sort=False)
    print(combined_df.index.is_unique)
    return combined_df

def fetch_with_retry(fetch_function, *args, max_retries=5, delay=5, **kwargs):
    retries = 0
    while retries < max_retries:
        try:
            result = fetch_function(*args, **kwargs)
            if result is not None and not result.empty:
                return result
        except Exception as e:
            print(f"Error fetching data: {e}")
        retries += 1
        print(f"Retrying {retries}/{max_retries}...")
        time.sleep(delay)
    raise RuntimeError("Failed to fetch data after several retries.")
